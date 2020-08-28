#导包
from openpyxl import load_workbook
from tools.get_path import *
from tools.read_config import ReadConfig
from tools.get_data import GetData

class DoExecl:
    #读取execl表格数据:参数{file_name:要打开的文件名；sheet_name:操作的工作簿名}
    @classmethod
    def read_execl(cls,file_name):
            # #打开execl文件
            wb=load_workbook(file_name)
            #存储读取的测试用例数据
            test_datas = []
            #获取配置文件数据
            mode=eval(ReadConfig.get_config(case_config_path,'MODE','mode'))
            #获取execl表单中的手机号
            tel=int(getattr(GetData, "tel"))
            # 借款的用户id
            memberId = int(getattr(GetData, "memberId"))
            # 借款用户手机号
            borrower_tel = int(getattr(GetData, "borrower_tel"))
            # 管理员手机号
            admin_tel = int(getattr(GetData, "admin_tel"))
            # 投资用户手机号
            investor_tel = int(getattr(GetData, "investor_tel"))
            # 投资用户id
            investor_id = int(getattr(GetData, "investor_id"))
            #遍历配置文件数据的key值
            for key in mode:
                # 获取工作簿
                sheet = wb[key]
                # br=sheet.max_row+1
                if mode[key]=='all':
                    for i in range(2,sheet.max_row+1):
                        row_data={}
                        row_data['case_id'] = sheet.cell(i, 1).value
                        row_data['title']=sheet.cell(i,2).value
                        row_data['url'] = sheet.cell(i,3).value
                        # row_data['data'] = sheet.cell(i,4).value
                        #手机号的数据替换
                        if sheet.cell(i,4).value.find('${tel}')!=-1:
                            row_data['data'] = sheet.cell(i, 4).value.replace('${tel}',str(tel))
                        elif sheet.cell(i,4).value.find('${borrower_tel}')!=-1:
                            row_data['data'] = sheet.cell(i, 4).value.replace('${borrower_tel}', str(borrower_tel))
                        elif sheet.cell(i,4).value.find('${memberId}')!=-1:
                            row_data['data'] = sheet.cell(i, 4).value.replace('${memberId}', str(memberId))
                        elif sheet.cell(i,4).value.find('${admin_tel}')!=-1:
                            row_data['data'] = sheet.cell(i, 4).value.replace('${admin_tel}', str(admin_tel))
                        elif sheet.cell(i,4).value.find('${investor_tel}')!=-1:
                            row_data['data'] = sheet.cell(i, 4).value.replace('${investor_tel}', str(investor_tel))
                        elif sheet.cell(i,4).value.find('${investor_id}')!=-1:
                            row_data['data'] = sheet.cell(i, 4).value.replace('${investor_id}', str(investor_id))
                        else:
                            row_data['data'] = sheet.cell(i, 4).value
                        row_data['http_method'] = sheet.cell(i,5).value
                        row_data['expected'] = sheet.cell(i, 6).value
                        row_data['mode']=key
                        test_datas.append(row_data)
                        cls.update_tel(file_name,'init',tel+1)
                else:
                    for i in mode[key]:
                        row_data = {}
                        row_data['case_id'] = sheet.cell(i+1, 1).value
                        row_data['title'] = sheet.cell(i+1, 2).value
                        row_data['url'] = sheet.cell(i+1, 3).value
                        # row_data['data'] = sheet.cell(i, 4).value
                        #手机号的数据替换
                        if sheet.cell(i+1,4).value.find('${tel}')!=-1:
                            row_data['data'] = sheet.cell(i+1, 4).value.replace('${tel}',str(tel))
                        elif sheet.cell(i+1,4).value.find('${borrower_tel}')!=-1:
                            row_data['data'] = sheet.cell(i+1, 4).value.replace('${borrower_tel}', str(borrower_tel))
                        elif sheet.cell(i+1,4).value.find('${memberId}')!=-1:
                            row_data['data'] = sheet.cell(i+1, 4).value.replace('${memberId}', str(memberId))
                        elif sheet.cell(i+1,4).value.find('${admin_tel}')!=-1:
                            row_data['data'] = sheet.cell(i+1, 4).value.replace('${admin_tel}', str(admin_tel))
                        elif sheet.cell(i+1,4).value.find('${investor_tel}')!=-1:
                            row_data['data'] = sheet.cell(i+1, 4).value.replace('${investor_tel}', str(investor_tel))
                        elif sheet.cell(i+1,4).value.find('${investor_id}')!=-1:
                            row_data['data'] = sheet.cell(i+1, 4).value.replace('${investor_id}', str(investor_id))
                        else:
                            row_data['data'] = sheet.cell(i+1, 4).value
                        row_data['http_method'] = sheet.cell(i+1, 5).value
                        row_data['expected'] = sheet.cell(i+1, 6).value
                        row_data['mode'] = key
                        test_datas.append(row_data)
                        cls.update_tel(file_name, 'init', tel + 1)
            #获取单元格数据,输入第一行，第一列的值
            #print(sheet.cell(1,1).value)
            return test_datas
    #数据写入execl，参数{file_name:要写入的文件名；sheet_name:操作的工作簿名，i：要写入的行数，vaule：要写入的数据}
    @staticmethod
    def write_execl(file_name,sheet_name,i,vaule,test_result=''):
        # 打开execl文件
        wb = load_workbook(file_name)
        # 获取工作簿
        sheet = wb[sheet_name]
        #
        sheet.cell(i,7).value=vaule
        sheet.cell(i,8).value = test_result
        #保存结果
        wb.save(file_name)

    @classmethod
    def update_tel(cls,file_name,sheet_name,tel):
        # 打开execl文件
        wb = load_workbook(file_name)
        # 获取工作簿
        sheet = wb[sheet_name]
        #第二行，第一列写入新的手机号
        sheet.cell(2,1).value = tel
        # 保存结果
        wb.save(file_name)

if __name__ == '__main__':
    print(DoExecl().read_execl('../test_data/login_testcase.xlsx'))